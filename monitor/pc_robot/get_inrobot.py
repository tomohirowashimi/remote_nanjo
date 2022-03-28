# // 20220217 新規作成
# // 0302 見直し？？
# // 0303 12用編集

import pandas as pd
import openpyxl as px
import csv
import datetime
import platform
from smb.SMBConnection import SMBConnection
import time
import os

path = os.getcwd()

f5_read_path = "59K).xls"
f6_read_path = ").xls"
f5_read_sheet = "59"
f6_read_sheet = "F6"
cell_read_sheet = "12"     # // 0303 追記
f5_write_day = "product_f5_day.csv"
f5_write_night = "product_f5_night.csv"
f6_write_day = "product_f6_day.csv"
f6_write_night = "product_f6_night.csv"
cell_write_day = "product_12_day.csv"     # // 0303 追記
cell_write_night = "product_12_night.csv"     # // 0303 追記
select_day = "day"
select_night = "night"

def make_plan(read_path, read_sheet, write_file, select):
    now = datetime.datetime.now()
    year = str(now.year)[2:]
    month = str(now.month)
    day = str(now.day)

    if select == "day":
        work_time = 'Ｄ勤'
    elif select == "night":
        work_time = 'Ｎ勤'
    cycle_data = "合計稼動時間"

    file_name = path + "\生産計画(’" + year + "_" + month + "_" + day + read_path
    print(file_name)
    sheet = read_sheet
    change_file_name = "temp" + ".xlsx"
    change_sheet = "Sheet1"

    df = pd.read_excel(file_name, sheet_name=sheet, header=None)
    df.to_excel(change_file_name, index=False, header=False)

    wb = px.load_workbook(change_file_name)
    ws = wb[change_sheet]
    
    qua = get_qua(work_time, ws)
    cycle = get_cycle(cycle_data, ws)
    list = []
    list.append(qua)
    list.append(cycle)

    with open(write_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(list)    

def get_qua(work_time, ws):
    qua = 0
    for row in ws["C10:C100"]:
        for cell in row:
            cell_value = cell.value
            o = cell.offset(0,1)
            offset = o.value

            if cell_value == work_time:
                print(cell.coordinate, cell_value, offset)

                try:
                    qua = qua + offset
                except TypeError:
                    print("No Shift")
                    
    return qua

def get_cycle(cycle_data, ws):
    for row in ws["I10:I100"]:
        for cell in row:
            cell_value = cell.value
            o = cell.offset(-5,0)
            offset = o.value

            if cell_value == cycle_data:
                print(cell.coordinate, cell_value, offset)
                cycle = offset
    return cycle

if __name__ == "__main__":
    make_plan(f5_read_path, f5_read_sheet, f5_write_day, select_day)
    make_plan(f5_read_path, f5_read_sheet, f5_write_night, select_night)
    make_plan(f6_read_path, f6_read_sheet, f6_write_day, select_day)
    make_plan(f6_read_path, f6_read_sheet, f6_write_night, select_night)
    make_plan(f6_read_path, cell_read_sheet, cell_write_day, select_day)
    make_plan(f6_read_path, cell_read_sheet, cell_write_night, select_night)
